import cv2
import time

import glob
import os
from openpyxl import Workbook
from ultralytics import YOLO
from PIL import Image
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import configparser

script_dir = os.path.dirname(os.path.abspath(__file__))
config = configparser.ConfigParser()
config.read('setting.ini')

target_dir = config['setting']['TARGET_DIR ']
save_dir = config['setting']['SAVE_DIR ']

day_files = sorted(glob.glob("/media/frog-takuma/imation i9/*"))
model = YOLO("model/yolov8x.pt") 
model.to("cuda:0")

GREEN = (0, 200, 0)
ORANGE = (0, 100, 255)
BLUE = (255, 100, 0)
conf_threshold = 0.90

all_data = {}

excel_filename = 'time_count_data.xlsx'

if not os.path.exists(excel_filename):
    # エクセルファイルが存在しない場合は新規作成
    workbook = Workbook()
    workbook.save(excel_filename)

try:
    existing_df = pd.read_excel(excel_filename)
except FileNotFoundError:
    existing_df = pd.DataFrame(columns=['year','month','day','hour','minute','Count'])

new_df = pd.DataFrame(columns=['year','month','day','hour','minute','Count'])

for df in day_files:
    hour_files = sorted(glob.glob(os.path.join(df,"*")))

    day_data = []
    sheet_name = df.split("/")[-1]


    for hf in hour_files:
        min_files = sorted(glob.glob(os.path.join(hf,"*")))
        print(len(min_files))

        for mf in min_files:
            
            print(mf)
            cap = cv2.VideoCapture(mf)

            if not cap.isOpened():
                print("Error: 動画ファイルを開くことができません。")
                exit()

            # フレームを一つずつ読み取る
            while cap.isOpened():
                ret, frame = cap.read()


                start_time = time.time()
                results = model(frame, conf=0.5, verbose=False)
                end_time = time.time()
                execution_time = end_time - start_time

                print("実行時間: {}秒".format(execution_time))

                detections = results[0].boxes.data
                detections = detections.to("cpu").numpy()
                class_0_count = np.sum(detections[:, -1] == 0)

                year = mf.split("/")[-3][:4]
                month = mf.split("/")[-3][4:6]
                day = mf.split("/")[-3][6:8]
                hour = mf.split("/")[-2]
                minute = mf.split("/")[-1][:2]
                second = "00"

                counts_over_threshold = 0
                counts_under_threshold = 0

                for x0, y0, x1, y1, conf, class_id in detections:
                    x0 = int(x0)
                    y0 = int(y0)
                    x1 = int(x1)
                    y1 = int(y1)
                    
                    if conf > conf_threshold:
                        color = GREEN
                        counts_over_threshold += 1
                    else:
                        color = ORANGE
                        counts_under_threshold += 1

                    if class_id == 0:
                        cv2.rectangle(frame,
                                    pt1 = (int(x0), int(y0)),
                                    pt2 = (int(x1), int(y1)),
                                    color = color,
                                    thickness = 2,
                                    lineType = cv2.LINE_AA)
                        
                        cv2.putText(frame,
                                    text = f"{conf:.2f}",
                                    org = (int(x0), int(y0)-5),
                                    fontFace = cv2.FONT_HERSHEY_SIMPLEX,
                                    fontScale = 0.8,
                                    color = color,
                                    thickness = 2,
                                    lineType = cv2.LINE_AA)

                save_dir = os.path.join("save_img", f"{year}{month}{day}")
                os.makedirs(save_dir, exist_ok=True)

                filename = f"{year}{month}{day}{hour}{minute}{second}_{class_0_count}.jpg"
                filepath = os.path.join(save_dir, filename)
                print(filepath)       

                frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                pil_img = Image.fromarray(frame_rgb)
                #pil_img.save(filepath)   

                new_data = {
                    'ymd': f"{year}/{month}/{day}",  
                    'hm': f"{hour}:{minute}",
                    'Count': class_0_count  
                }
                """
                new_data = {
                    'ym': year,  
                    '': month,  
                    'day': day,  
                    'hour': hour,
                    'minute': minute,    
                    'Count': class_0_count  
                }
                """
                day_data.append(new_data)


                break
        
    new_df = pd.DataFrame(day_data)

    try:
        book = load_workbook(excel_filename)
        if sheet_name in book.sheetnames:
            # 既存のシートを読み込む
            existing_df = pd.read_excel(excel_filename, sheet_name=sheet_name)
            # 既存のデータに新しいデータを追加
            updated_df = pd.concat([existing_df, new_df], ignore_index=True)
        else:
            # 新しいシートを作成
            updated_df = new_df
    except FileNotFoundError:
        # ファイルが存在しない場合、新しいデータフレームをそのまま使用
        updated_df = new_df

    with pd.ExcelWriter(excel_filename, engine='openpyxl', mode='a' if 'book' in locals() else 'w', if_sheet_exists='replace') as writer:
        updated_df.to_excel(writer, sheet_name=sheet_name, index=False)        
